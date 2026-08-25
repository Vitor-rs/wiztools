#!/usr/bin/env python3
"""exportar-modelo-xlsx.py — o banco inteiro em planilha: uma aba por tabela.

    pip install openpyxl
    python exportar-modelo-xlsx.py wizard-ensaio.db modelo-dados-aluno.xlsx

Ferramenta de DESENVOLVIMENTO, não faz parte do app — por isso pode depender de Python e
openpyxl sem quebrar a promessa de zero dependências do `main.ts`. Aponte-a para um banco de
ENSAIO: o `modelo-dados-aluno.xlsx` que vive no repositório saiu do banco do aluno fictício, e
gerá-lo a partir do `wizard.db` publicaria os nomes reais da escola num repositório público.

Existe porque o `schema.sql` só descreve 18 das 42 tabelas — as outras 24 nascem das migrações
do `main.ts`. Sem isto não há onde ler o modelo de dados inteiro de uma vez.

Refaça a planilha quando o esquema mudar: tabela nova aparece sozinha, mas precisa ser posta
num dos grupos de GRUPOS abaixo, senão o script avisa que ela ficou de fora do índice.
"""
import sqlite3, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB, OUT = sys.argv[1], sys.argv[2]

# agrupadas por papel — é assim que o índice explica o sistema
GRUPOS = [
    ("Cadastro do aluno", ["alunos", "aluno_livro", "aluno_estagio",
                           "aluno_situacao_historico", "aluno_horario_historico"]),
    ("Agenda e frequência", ["aulas", "aula_professor", "presenca", "encontro_avulso", "diario"]),
    ("Turmas", ["turmas", "turma_dia", "turma_professor"]),
    ("Catálogo pedagógico", ["estagio", "estagio_licao", "estagio_licao_extra", "estagio_modelo",
                             "estagio_modelo_licao", "estagio_proximo", "estagio_equivalente",
                             "estagio_anexo"]),
    ("Estoque e material", ["estoque_item", "estoque_edicao", "estoque_unidade", "estoque_kit_item",
                            "estoque_evento", "estoque_evento_item", "entrega_material",
                            "devolucao_material"]),
    ("Calendário letivo", ["calendario_marcacao", "calendario_trecho", "calendario_fonte",
                           "calendario_importado", "calendario_sync"]),
    ("Domínio e ajustes", ["livros", "dias", "situacoes", "prioridade", "horario_ativo",
                           "funcionarios", "config", "aviso_silenciado"]),
]

CAB = PatternFill("solid", fgColor="1F3864")
CAB_F = Font(color="FFFFFF", bold=True, size=10)
TIT = Font(bold=True, size=13, color="1F3864")
SUB = Font(italic=True, size=9, color="666666")

c = sqlite3.connect(DB)
c.row_factory = sqlite3.Row
existentes = {r[0] for r in c.execute(
    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")}

wb = Workbook()
ix = wb.active
ix.title = "ÍNDICE"
ix["A1"] = "Wizard Naviraí — modelo de dados"
ix["A1"].font = Font(bold=True, size=15, color="1F3864")
ix["A2"] = "Uma aba por tabela do banco. Os dados são de um único aluno-modelo: João da Silva (9001), Kids 2 3rd Edition."
ix["A2"].font = SUB
linha = 4
for cel, t in (("A", "Tabela"), ("B", "Linhas"), ("C", "Colunas"), ("D", "O que guarda")):
    ix[f"{cel}{linha}"] = t
    ix[f"{cel}{linha}"].fill = CAB
    ix[f"{cel}{linha}"].font = CAB_F
linha += 1

vistas = set()
for grupo, tabelas in GRUPOS:
    ix[f"A{linha}"] = grupo
    ix[f"A{linha}"].font = TIT
    linha += 1
    for t in tabelas:
        if t not in existentes:
            continue
        vistas.add(t)
        cols = [r[1] for r in c.execute(f"PRAGMA table_info({t})")]
        linhas = list(c.execute(f"SELECT * FROM {t}"))
        ws = wb.create_sheet(t[:31])
        for j, col in enumerate(cols, 1):
            cel = ws.cell(1, j, col)
            cel.fill = CAB
            cel.font = CAB_F
            cel.alignment = Alignment(horizontal="center")
        for i, r in enumerate(linhas, 2):
            for j, col in enumerate(cols, 1):
                ws.cell(i, j, r[col])
        ws.freeze_panes = "A2"
        if linhas:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(linhas)+1}"
        for j, col in enumerate(cols, 1):
            largura = max(len(col), *(len(str(r[col])) if r[col] is not None else 0
                                      for r in linhas)) if linhas else len(col)
            ws.column_dimensions[get_column_letter(j)].width = min(max(largura + 2, 9), 42)
        ix[f"A{linha}"] = t
        ix[f"A{linha}"].hyperlink = f"#'{t[:31]}'!A1"
        ix[f"A{linha}"].font = Font(color="0563C1", underline="single")
        ix[f"B{linha}"] = len(linhas)
        ix[f"C{linha}"] = len(cols)
        ix[f"D{linha}"] = ", ".join(cols[:6]) + (" …" if len(cols) > 6 else "")
        ix[f"D{linha}"].font = SUB
        linha += 1
    linha += 1

faltando = existentes - vistas
if faltando:
    print("!! tabelas fora dos grupos:", sorted(faltando))

for col, w in (("A", 30), ("B", 9), ("C", 9), ("D", 90)):
    ix.column_dimensions[col].width = w
ix.freeze_panes = "A5"
wb.save(OUT)
print(f"{OUT}: {len(wb.sheetnames)-1} abas de tabela + índice")
